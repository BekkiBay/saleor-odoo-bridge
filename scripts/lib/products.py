"""Импорт product.template из xlsx.

Колонки (порядок строго как в test-catalog-clothing-v2.xlsx):
  Артикул | Название | Категория товаров | Цена продажи (UZS) |
  Себестоимость (UZS) | Штрихкод | Описание
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import odoorpc
from openpyxl import load_workbook


COL_SKU = "Артикул"
COL_NAME = "Название"
COL_CATEGORY = "Категория товаров"
COL_PRICE = "Цена продажи (UZS)"
COL_COST = "Себестоимость (UZS)"
COL_BARCODE = "Штрихкод"
COL_DESC = "Описание"


@dataclass
class ProductRow:
    sku: str
    name: str
    category_path: str
    list_price: float
    standard_price: float
    barcode: str
    description: str


def read_catalog(path: str | Path) -> list[ProductRow]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    expected = [COL_SKU, COL_NAME, COL_CATEGORY, COL_PRICE, COL_COST, COL_BARCODE, COL_DESC]
    missing = [c for c in expected if c not in header]
    if missing:
        raise RuntimeError(
            f"в xlsx не хватает колонок: {missing}. Найдены: {header}"
        )
    idx = {c: header.index(c) for c in expected}

    rows: list[ProductRow] = []
    for raw in rows_iter:
        if raw is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in raw):
            continue
        rows.append(
            ProductRow(
                sku=str(raw[idx[COL_SKU]]).strip(),
                name=str(raw[idx[COL_NAME]]).strip(),
                category_path=str(raw[idx[COL_CATEGORY]]).strip(),
                list_price=float(raw[idx[COL_PRICE]] or 0),
                standard_price=float(raw[idx[COL_COST]] or 0),
                barcode=str(raw[idx[COL_BARCODE]]).strip() if raw[idx[COL_BARCODE]] is not None else "",
                description=str(raw[idx[COL_DESC]]).strip() if raw[idx[COL_DESC]] is not None else "",
            )
        )
    return rows


def _storable_field(odoo: odoorpc.ODOO) -> tuple[str, object]:
    """Какое поле в product.template отвечает за 'хранимый товар' в этой версии?

    Odoo 17+: is_storable (bool).
    Odoo <17: type='product'.
    """
    Tmpl = odoo.env["ir.model.fields"]
    ids = Tmpl.search([("model", "=", "product.template"), ("name", "=", "is_storable")])
    if ids:
        return "is_storable", True
    return "type", "product"


def import_products(odoo: odoorpc.ODOO, rows: list[ProductRow],
                    category_mapping: dict[str, int]) -> dict[str, int]:
    """Создать/обновить product.template. Возвращает {'created': N, 'updated': M}."""
    Tmpl = odoo.env["product.template"]
    storable_field, storable_value = _storable_field(odoo)
    print(f"  storable field detected: {storable_field}={storable_value}")

    created = updated = 0
    for r in rows:
        if r.category_path not in category_mapping:
            raise RuntimeError(f"для SKU {r.sku} не найден category_id (path={r.category_path!r})")

        vals = {
            "default_code": r.sku,
            "name": r.name,
            "categ_id": category_mapping[r.category_path],
            "list_price": r.list_price,
            "standard_price": r.standard_price,
            "barcode": r.barcode or False,
            "description_sale": r.description or False,
            "active": True,
            storable_field: storable_value,
        }

        existing = Tmpl.search([("default_code", "=", r.sku)])
        if existing:
            Tmpl.browse(existing[0]).write(vals)
            updated += 1
        else:
            Tmpl.create(vals)
            created += 1

    print(f"  → created={created}, updated={updated}")
    return {"created": created, "updated": updated}
